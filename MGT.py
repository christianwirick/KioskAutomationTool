"""
Process a main Excel file using MGT prefix lookup data, then export results.

Workflow:
1. User selects the main workbook and lookup workbook.
2. Script inserts derived columns into the main workbook data.
3. Script builds a list of output values grouped by segment.
4. User chooses one combined export or one file per segment.
"""

from collections import defaultdict
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, simpledialog

import openpyxl


OUTPUT_DIR = Path("MGT")


class ExcelProcessor:
    """Handle loading, transforming, and exporting Excel data."""

    def __init__(self, main_file: str, lookup_file: str) -> None:
        self.main_file = main_file
        self.lookup_file = lookup_file
        self.main_wb = None
        self.lookup_wb = None
        self.segment_data = defaultdict(list)
        self.mgt_data = {}

    def load_workbooks(self) -> None:
        """Load both workbooks into memory."""
        try:
            self.main_wb = openpyxl.load_workbook(self.main_file)
            self.lookup_wb = openpyxl.load_workbook(self.lookup_file)
        except Exception as exc:
            print(f"Error loading workbooks: {exc}")
            raise

    def prepare_lookup_data(self) -> None:
        """
        Build a dictionary from the lookup sheet.

        Expected structure (starting row 2):
        - Column A: prefix code
        - Column B: mapped MGT value
        """
        try:
            lookup_sheet = self.lookup_wb.active
            self.mgt_data = {
                row[0].value: row[1].value
                for row in lookup_sheet.iter_rows(min_row=2)
                if row[0].value is not None
            }
        except Exception as exc:
            print(f"Error preparing lookup data: {exc}")
            raise

    def process_data(self) -> None:
        """
        Transform rows in the main workbook and collect export values by segment.

        Derived columns inserted at D:G:
        - D = left 2 chars of column C
        - E = lookup value from prefix
        - F = right 9 chars of column C
        - G = concatenation of E and F
        """
        try:
            sheet = self.main_wb.active

            # Insert 4 empty columns starting at D so derived data has dedicated columns.
            sheet.insert_cols(4, 4)

            for row in range(2, sheet.max_row + 1):
                raw_value = sheet[f"C{row}"].value

                # Convert to string before slicing so numeric cells are handled safely.
                code_text = str(raw_value) if raw_value is not None else ""
                left_chars = code_text[:2]
                right_chars = code_text[-9:]

                sheet[f"D{row}"] = left_chars
                sheet[f"E{row}"] = self.mgt_data.get(left_chars, "")
                sheet[f"F{row}"] = right_chars
                sheet[f"G{row}"] = f"{sheet[f'E{row}'].value}{sheet[f'F{row}'].value}"

                # Segment name expected in column H based on original sheet design.
                segment_name = sheet[f"H{row}"].value
                if segment_name:
                    self.segment_data[segment_name].append(sheet[f"G{row}"].value)

            print(f"Processed {sheet.max_row - 1} rows.")
        except Exception as exc:
            print(f"Error processing data: {exc}")
            raise

    def export_data(self) -> None:
        """Ask user how to export and route to the chosen method."""
        choice = simpledialog.askstring(
            "Choose Export Option",
            "Type 'ALL' for a single file or 'BY SEGMENT' for separate files",
        )

        if not choice:
            print("Export canceled.")
            return

        choice = choice.strip().upper()
        if choice == "ALL":
            self.export_all_data()
        elif choice == "BY SEGMENT":
            self.export_by_segment()
        else:
            print("Invalid choice. Please type 'ALL' or 'BY SEGMENT'.")

    @staticmethod
    def _to_integer(value):
        """
        Convert value to integer for export.

        The original logic was int(float(value)); kept for compatibility with
        numeric strings and float-like values.
        """
        return int(float(value))

    def export_all_data(self) -> None:
        """Export all segment values into a single file."""
        try:
            OUTPUT_DIR.mkdir(exist_ok=True)

            all_data_wb = openpyxl.Workbook()
            all_data_sheet = all_data_wb.active

            row = 1
            for data in self.segment_data.values():
                for item in data:
                    all_data_sheet[f"A{row}"].value = self._to_integer(item)
                    # Force integer display (no decimals / scientific notation).
                    all_data_sheet[f"A{row}"].number_format = "0"
                    row += 1

            all_data_filename = OUTPUT_DIR / "All_Data.xlsx"
            all_data_wb.save(all_data_filename)
            print("Exported all data to a single Excel file.")
        except Exception as exc:
            print(f"Error exporting all data: {exc}")
            raise

    def export_by_segment(self) -> None:
        """Export each segment's values into its own file."""
        try:
            OUTPUT_DIR.mkdir(exist_ok=True)

            for segment, data in self.segment_data.items():
                segment_wb = openpyxl.Workbook()
                segment_sheet = segment_wb.active

                row = 1
                for item in data:
                    segment_sheet[f"A{row}"].value = self._to_integer(item)
                    # Keep values formatted as whole numbers in Excel.
                    segment_sheet[f"A{row}"].number_format = "0"
                    row += 1

                segment_filename = OUTPUT_DIR / f"{segment}.xlsx"
                segment_wb.save(segment_filename)

            print("Exported data by segment into individual Excel files.")
        except Exception as exc:
            print(f"Error exporting by segment: {exc}")
            raise


def select_excel_file(title: str) -> str:
    """Open a file picker and return selected Excel file path."""
    root = tk.Tk()
    root.withdraw()

    filename = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
    )

    root.destroy()
    return filename


def main() -> None:
    """Entry point for interactive file selection and processing."""
    main_file = select_excel_file("Select the main Excel file")
    if not main_file:
        print("No main file selected.")
        return

    lookup_file = select_excel_file("Select the MGT Prefix Codes Excel file")
    if not lookup_file:
        print("No lookup file selected.")
        return

    processor = ExcelProcessor(main_file, lookup_file)
    processor.load_workbooks()
    processor.prepare_lookup_data()
    processor.process_data()
    processor.export_data()

    print("Processing completed successfully.")


if __name__ == "__main__":
    main()
